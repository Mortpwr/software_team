import csv
from io import BytesIO, StringIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.deps import CurrentSession, get_current_session
from app.models import Student
from app.services.common import audit
from app.services.permissions import COORDINATOR, LEADER, TEACHER, require_roles, scoped_student_ids
from app.services.serializers import student_public

router = APIRouter(tags=["students"])

FIELD_ALIASES = {
    "studentId": ["studentId", "student_id", "学号"],
    "name": ["name", "姓名"],
    "grade": ["grade", "年级"],
    "major": ["major", "专业"],
    "className": ["className", "class_name", "班级"],
    "nation": ["nation", "民族"],
    "phone": ["phone", "手机号", "联系方式"],
    "politicalStatus": ["politicalStatus", "political_status", "政治面貌"],
    "tutor": ["tutor", "导师"],
    "hometown": ["hometown", "生源地", "户籍地"],
}
REQUIRED_FIELDS = ["studentId", "name", "grade", "major", "className"]


@router.get("/student/me")
def me(db: Session = Depends(get_db), session: CurrentSession = Depends(get_current_session)) -> dict:
    student = db.get(Student, session.student_id)
    if not student:
        raise HTTPException(status_code=404, detail="student not found")
    return student_public(student, session.role)


@router.get("/students")
def list_students(db: Session = Depends(get_db), session: CurrentSession = Depends(get_current_session)) -> dict:
    require_roles(session, TEACHER, LEADER, COORDINATOR)
    stmt = select(Student)
    scope = scoped_student_ids(db, session)
    if scope is not None:
        stmt = stmt.where(Student.student_id.in_(scope))
    rows = db.scalars(stmt.order_by(Student.grade.desc(), Student.student_id)).all()
    return {"list": [student_public(row, session.role) for row in rows]}


@router.get("/students/export")
def export_students(db: Session = Depends(get_db), session: CurrentSession = Depends(get_current_session)) -> Response:
    require_roles(session, TEACHER)
    rows = db.scalars(select(Student).order_by(Student.grade.desc(), Student.student_id)).all()
    fp = StringIO()
    writer = csv.writer(fp)
    writer.writerow(["学号", "姓名", "年级", "专业", "班级", "民族", "手机号(脱敏)", "政治面貌", "导师"])
    for row in rows:
        item = student_public(row, "leader")
        writer.writerow([
            item["studentId"],
            item["name"],
            item["grade"],
            item["major"],
            item["className"],
            item["nation"],
            item["phoneMasked"],
            item["politicalStatus"],
            item["tutor"],
        ])
    filename = quote("学生画像导出.csv")
    return Response(
        "\ufeff" + fp.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/students/import")
async def import_students(
    file: UploadFile = File(...),
    dry_run: bool = Form(default=True, alias="dryRun"),
    overwrite: bool = Form(default=False),
    db: Session = Depends(get_db),
    session: CurrentSession = Depends(get_current_session),
) -> dict:
    require_roles(session, TEACHER)
    rows = await parse_student_rows(file)
    result = validate_student_rows(db, rows, overwrite)
    if dry_run or result["errors"]:
        return {"ok": not result["errors"], "dryRun": True, **result}

    created = 0
    updated = 0
    for item in result["validRows"]:
        existing = db.get(Student, item["studentId"])
        if existing:
            apply_student(existing, item)
            updated += 1
        else:
            db.add(
                Student(
                    student_id=item["studentId"],
                    name=item["name"],
                    grade=item["grade"],
                    major=item["major"],
                    class_name=item["className"],
                    nation=item.get("nation", ""),
                    phone=item.get("phone", ""),
                    political_status=item.get("politicalStatus", ""),
                    tutor=item.get("tutor", ""),
                    hometown=item.get("hometown", ""),
                    extension={},
                ),
            )
            created += 1
    audit(db, session, "students_import", file.filename or "students", {"created": created, "updated": updated})
    db.commit()
    return {"ok": True, "dryRun": False, "total": result["total"], "created": created, "updated": updated, "errors": [], "preview": result["preview"]}


async def parse_student_rows(file: UploadFile) -> list[dict]:
    raw = await file.read()
    suffix = (file.filename or "").lower()
    if suffix.endswith(".xlsx"):
        return parse_xlsx(raw)
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [{"row": index + 2, "data": normalize_row(row)} for index, row in enumerate(reader)]


def parse_xlsx(raw: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=400, detail="xlsx import requires openpyxl") from exc
    workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(item or "").strip() for item in rows[0]]
    parsed = []
    for index, values in enumerate(rows[1:], start=2):
        raw_row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        parsed.append({"row": index, "data": normalize_row(raw_row)})
    return parsed


def normalize_row(row: dict) -> dict:
    normalized = {}
    for target, aliases in FIELD_ALIASES.items():
        value = next((row.get(alias) for alias in aliases if row.get(alias) not in {None, ""}), "")
        normalized[target] = str(value).strip()
    return normalized


def validate_student_rows(db: Session, rows: list[dict], overwrite: bool) -> dict:
    errors = []
    valid_rows = []
    seen = set()
    for row in rows:
        line = row["row"]
        data = row["data"]
        missing = [field for field in REQUIRED_FIELDS if not data.get(field)]
        if missing:
            errors.append({"row": line, "field": ",".join(missing), "message": "必填字段缺失"})
            continue
        if data["studentId"] in seen:
            errors.append({"row": line, "field": "studentId", "message": "导入文件内学号重复"})
            continue
        seen.add(data["studentId"])
        if db.get(Student, data["studentId"]) and not overwrite:
            errors.append({"row": line, "field": "studentId", "message": "学号已存在，需勾选覆盖更新"})
            continue
        valid_rows.append(data)
    existing_ids = {item["studentId"] for item in valid_rows if db.get(Student, item["studentId"])}
    return {
        "total": len(rows),
        "created": sum(1 for item in valid_rows if item["studentId"] not in existing_ids),
        "updated": sum(1 for item in valid_rows if item["studentId"] in existing_ids),
        "errors": errors,
        "validRows": valid_rows,
        "preview": valid_rows[:5],
    }


def apply_student(row: Student, data: dict) -> None:
    row.name = data["name"]
    row.grade = data["grade"]
    row.major = data["major"]
    row.class_name = data["className"]
    row.nation = data.get("nation", row.nation)
    row.phone = data.get("phone", row.phone)
    row.political_status = data.get("politicalStatus", row.political_status)
    row.tutor = data.get("tutor", row.tutor)
    row.hometown = data.get("hometown", row.hometown)
